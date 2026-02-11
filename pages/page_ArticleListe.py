import customtkinter
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import psycopg2
import json
import subprocess
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

# ====================================================================
# GESTION DES IMPORTATIONS (Compatible VSCode ET app_main)
# ====================================================================

def import_page_info_article():
    """Tente d'importer PageInfoArticle avec différentes stratégies"""
    try:
        # Essai 1 : Import depuis pages.page_infoArticle (pour app_main)
        from pages.page_infoArticle import PageInfoArticle
        print("✓ PageInfoArticle importée depuis pages.page_infoArticle")
        return PageInfoArticle
    except ImportError:
        try:
            # Essai 2 : Import direct depuis page_infoArticle (pour VSCode/standalone)
            from page_infoArticle import PageInfoArticle
            print("✓ PageInfoArticle importée depuis page_infoArticle")
            return PageInfoArticle
        except ImportError as e:
            print(f"❌ Erreur d'import PageInfoArticle: {e}")
            
            # Classe de substitution si aucun import ne fonctionne
            class PageInfoArticleFallback(customtkinter.CTkFrame):
                def __init__(self, master, db_conn=None, session_data=None, initial_idarticle=None):
                    super().__init__(master)
                    self.pack(fill="both", expand=True)
                    
                    error_frame = customtkinter.CTkFrame(self, fg_color="#ffebee")
                    error_frame.pack(fill="both", expand=True, padx=20, pady=20)
                    
                    customtkinter.CTkLabel(
                        error_frame,
                        text="❌ Erreur de chargement",
                        font=customtkinter.CTkFont(size=20, weight="bold"),
                        text_color="#c62828"
                    ).pack(pady=20)
                    
                    customtkinter.CTkLabel(
                        error_frame,
                        text=f"Impossible de charger PageInfoArticle\nArticle ID: {initial_idarticle}",
                        font=customtkinter.CTkFont(size=12),
                        text_color="#666"
                    ).pack(pady=10)
            
            return PageInfoArticleFallback

# Importer PageInfoArticle avec la fonction adaptative
PageInfoArticle = import_page_info_article()

# ====================================================================
# 1. Configuration du Style du Treeview (CORRIGÉE)
# ====================================================================

def configure_treeview_style(root):
    """
    Applique un style au Treeview (ttk) pour qu'il corresponde au thème CTK
    de manière simplifiée pour éviter les bugs de couleur.
    """
    style = ttk.Style(root)
    
    # Utiliser le thème 'default' pour pouvoir le configurer
    style.theme_use('default')
    
    # Couleurs stables (fonctionne bien avec le mode sombre CTK)
    bg_color = "#2b2b2b"      # Couleur de fond sombre pour le corps du Treeview
    text_color = "#FFFFFF"    # Couleur du texte blanc
    heading_bg = "#3e3e3e"    # Couleur de fond des en-têtes
    selected_bg = "#1f6aa5"   # Couleur de sélection (bleu CTK)
    
    # Configuration générale du Treeview
    style.configure("Treeview", 
                    background=bg_color,
                    foreground=text_color,
                    fieldbackground=bg_color,
                    borderwidth=0,
                    rowheight=25)
    
    # Configuration des en-têtes
    style.configure("Treeview.Heading",
                    background=heading_bg,
                    foreground=text_color,
                    font=('Arial', 10, 'bold'))
    
    # Configuration de la couleur de sélection
    style.map('Treeview', 
              background=[('selected', selected_bg)],
              foreground=[('selected', '#FFFFFF')])
    
    root.option_add('*Treeview*highlightThickness', 0)
    root.option_add('*Treeview*Font', 'Arial 10')


# ====================================================================
# 2. Classe de la Page d'Affichage (CTkFrame)
# ====================================================================

class page_listeArticle(customtkinter.CTkFrame):
    def __init__(self, master, db_conn=None, session_data=None, **kwargs):
        """
        MODIFICATION: Accepte db_conn et session_data au lieu de data_fetcher_func
        pour être compatible avec app_main.py
        """
        super().__init__(master, **kwargs)
        
        # Stocker la connexion DB
        self.db_conn = db_conn
        self.session_data = session_data
        
        # Appliquer le style au Treeview
        try:
            # L'appel à la fonction corrigée ci-dessus
            configure_treeview_style(master)
        except Exception as e:
            # Utile pour le débogage si le style échoue toujours
            print(f"Erreur d'application du style Treeview: {e}")
            pass

        # Configuration du layout
        self.grid_rowconfigure(2, weight=1)
        self.grid_columnconfigure(0, weight=1)
        
        self.all_data = []  # Pour stocker toutes les données
        
        # ====== Zone de Recherche Unique ======
        self.create_search_frame()
        
        # ====== Bouton Export Excel ======
        self.create_export_button()
        
        # ====== Treeview ======
        self.create_treeview()
        
        # ====== Label compteur en bas ======
        self.label_count = customtkinter.CTkLabel(
            self, 
            text="Nombre d'articles : 0", 
            font=customtkinter.CTkFont(size=12, weight="bold")
        )
        self.label_count.grid(row=3, column=0, padx=10, pady=(5, 10), sticky="w")
        
        # Chargement initial des données
        self.load_data()

    def connect_db(self):
        """Connexion à la base de données PostgreSQL"""
        # Si une connexion existe déjà, l'utiliser
        if self.db_conn:
            return self.db_conn
            
        # Sinon, créer une nouvelle connexion
        conn = None
        try:
            with open('config.json') as f:
                config = json.load(f)
                db_config = config['database']

            conn = psycopg2.connect(
                host=db_config['host'],
                user=db_config['user'],
                password=db_config['password'],
                database=db_config['database'],
                port=db_config['port']
            )
            return conn
        except FileNotFoundError:
            messagebox.showerror("Erreur de configuration", "Fichier 'config.json' non trouvé.")
        except KeyError:
            messagebox.showerror("Erreur de configuration", "Clés de base de données manquantes dans 'config.json'.")
        except psycopg2.Error as err:
            messagebox.showerror("Erreur de connexion", f"Erreur de connexion à PostgreSQL : {err}")
        except Exception as err:
             messagebox.showerror("Erreur Inattendue", f"Une erreur inattendue est survenue : {err}")
        
        return None

    def fetch_articles_from_db(self):
        """
        Exécute la requête SQL de jointure et retourne les résultats.
        """
        conn = self.connect_db()
        if conn is None:
            return []

        # Requête SQL
        SQL_QUERY = """
        SELECT
            T2."idarticle",
            T1."codearticle",
            T2."designation",
            T1."designationunite",
            T1."qtunite",
            T1."poids",
            T3."designationcat"
        FROM
            tb_unite AS T1
        INNER JOIN
            tb_article AS T2 ON T1.idarticle = T2.idarticle
        INNER JOIN
            tb_categoriearticle AS T3 ON T2.idca = T3.idca
        ORDER BY T1."codearticle";
        """
        
        data = []
        try:
            with conn.cursor() as cur:
                cur.execute(SQL_QUERY)
                data = cur.fetchall()
            
        except psycopg2.Error as err:
            messagebox.showerror("Erreur SQL", f"Erreur lors de l'exécution de la requête : {err}")
            
        finally:
            # Ne fermer la connexion que si elle a été créée ici
            if conn and conn != self.db_conn:
                conn.close()
                
        return data

    def create_search_frame(self):
        """Crée le cadre de recherche avec une zone unique"""
        search_frame = customtkinter.CTkFrame(self)
        search_frame.grid(row=0, column=0, sticky='ew', padx=10, pady=(10, 5))
        search_frame.grid_columnconfigure(0, weight=1)
        
        # Label d'instruction
        label_search = customtkinter.CTkLabel(
            search_frame, 
            text="🔍 Recherche globale (cherche dans Code, Désignation, Unité et Catégorie):",
            font=customtkinter.CTkFont(size=12, weight="bold")
        )
        label_search.grid(row=0, column=0, padx=5, pady=(5, 2), sticky="w")
        
        # Frame pour la barre de recherche et le bouton
        search_input_frame = customtkinter.CTkFrame(search_frame, fg_color="transparent")
        search_input_frame.grid(row=1, column=0, sticky='ew', padx=5, pady=(0, 5))
        search_input_frame.grid_columnconfigure(0, weight=1)
        
        # Champ de recherche unique
        self.entry_search = customtkinter.CTkEntry(
            search_input_frame, 
            placeholder_text="Tapez votre recherche ici...",
            height=35,
            font=customtkinter.CTkFont(size=13)
        )
        self.entry_search.grid(row=0, column=0, padx=(0, 5), pady=0, sticky="ew")
        self.entry_search.bind('<KeyRelease>', lambda e: self.filter_data())
        
        # Bouton Réinitialiser
        btn_reset = customtkinter.CTkButton(
            search_input_frame, 
            text="✕ Effacer", 
            command=self.reset_filters,
            width=100,
            height=35,
            font=customtkinter.CTkFont(size=12)
        )
        btn_reset.grid(row=0, column=1, padx=0, pady=0)

    def create_export_button(self):
        """Crée les boutons d'action"""
        buttons_frame = customtkinter.CTkFrame(self)
        buttons_frame.grid(row=1, column=0, sticky='ew', padx=10, pady=(0, 5))
        
        # Bouton Nouvel Article (à gauche)
        btn_new_article = customtkinter.CTkButton(
            buttons_frame,
            text="➕ Nouvel Article",
            command=self.open_new_article,
            width=180,
            height=35,
            font=customtkinter.CTkFont(size=13, weight="bold"),
            fg_color="#007bff",
            hover_color="#0056b3"
        )
        btn_new_article.pack(side="left", padx=5, pady=5)
        
        # Bouton Nouvelle Catégorie
        btn_new_category = customtkinter.CTkButton(
            buttons_frame,
            text="📁 Nouvelle Catégorie",
            command=self.open_new_category,
            width=180,
            height=35,
            font=customtkinter.CTkFont(size=13, weight="bold"),
            fg_color="#6c757d",
            hover_color="#5a6268"
        )
        btn_new_category.pack(side="left", padx=5, pady=5)
        
        # Bouton Export Excel (à droite)
        btn_export = customtkinter.CTkButton(
            buttons_frame,
            text="📊 Exporter vers Excel",
            command=self.export_to_excel,
            width=180,
            height=35,
            font=customtkinter.CTkFont(size=13, weight="bold"),
            fg_color="#28a745",
            hover_color="#218838"
        )
        btn_export.pack(side="right", padx=5, pady=5)

    def create_treeview(self):
        """Crée le Treeview avec scrollbar"""
        # Frame pour le Treeview et scrollbar
        tree_frame = customtkinter.CTkFrame(self)
        tree_frame.grid(row=2, column=0, sticky='nsew', padx=10, pady=5)
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # Colonnes
        columns = ("ID", "Code", "Designation", "Unite", "Quantite", "Poids", "Categorie")
        
        # Création du Treeview
        self.tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)
        
        # Définition des en-têtes
        self.tree.heading("ID", text="ID")
        self.tree.heading("Code", text="Code Article")
        self.tree.heading("Designation", text="Désignation Article")
        self.tree.heading("Unite", text="Unité")
        self.tree.heading("Quantite", text="Quantité")
        self.tree.heading("Poids", text="Poids")
        self.tree.heading("Categorie", text="Catégorie")
        
        # Configuration des colonnes
        self.tree.column("ID", width=0, stretch=False)  # Colonne cachée
        self.tree.column("Code", width=120, anchor='center')
        self.tree.column("Designation", width=300, anchor='w')
        self.tree.column("Unite", width=80, anchor='w')
        self.tree.column("Quantite", width=100, anchor='e')
        self.tree.column("Poids", width=100, anchor='e')
        self.tree.column("Categorie", width=150, anchor='w')

        # Barre de défilement
        scrollbar = customtkinter.CTkScrollbar(tree_frame, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        # Disposition
        self.tree.grid(row=0, column=0, sticky='nsew')
        scrollbar.grid(row=0, column=1, sticky='ns')
        
        # Bind events
        self.tree.bind('<Double-Button-1>', self.on_double_click)
        self.tree.bind('<ButtonRelease-1>', self.on_single_click)

    def on_single_click(self, event):
        """Gère le clic simple"""
        selection = self.tree.selection()
        if selection:
            self.tree.selection_set(selection)

    def on_double_click(self, event):
        """Gère le double-clic sur une ligne"""
        selection = self.tree.selection()
        if selection:
            item = self.tree.item(selection[0])
            values = item['values']
            if values and len(values) > 0:
                idarticle = values[0]
                self.open_info_article(idarticle)

    def open_info_article(self, idarticle):
        """Ouvre la page d'information de l'article"""
        try:
            # PageInfoArticle est déjà importée au début du fichier
            info_window = customtkinter.CTkToplevel(self)
            info_window.title(f"Détails Article - ID: {idarticle}")
            info_window.geometry("1200x700")
    
            info_window.grid_columnconfigure(0, weight=1)
            info_window.grid_rowconfigure(0, weight=1)
    
            # Créer la page avec tous les arguments nommés
            page_frame = PageInfoArticle(
                master=info_window,
                db_conn=self.db_conn,
                session_data=self.session_data,
                initial_idarticle=str(idarticle)
            )
    
            page_frame.grid(row=0, column=0, sticky="nsew")
    
            # Centrer la fenêtre
            info_window.update_idletasks()
            x = (info_window.winfo_screenwidth() // 2) - (1200 // 2)
            y = (info_window.winfo_screenheight() // 2) - (700 // 2)
            info_window.geometry(f"1200x700+{x}+{y}")
    
            info_window.focus()
            info_window.lift()
    
        except Exception as e:
            messagebox.showerror(
                "Erreur", 
                f"Impossible d'ouvrir la page d'information:\n\n{str(e)}\n\nVérifiez que page_infoArticle.py est accessible."
            )
            import traceback
            traceback.print_exc()
            
    def open_new_article(self):
        """Ouvre la page de création d'un nouvel article"""
        try:
            from pages.page_article import PageArticle
            
            article_window = customtkinter.CTkToplevel(self)
            article_window.title("Nouvel Article")
            article_window.geometry("700x700")
            
            article_window.transient(self.master) 
            article_window.grab_set()
            
            article_window.grid_columnconfigure(0, weight=1)
            article_window.grid_rowconfigure(0, weight=1)
            
            # Gestion des arguments pour la compatibilité
            try:
                page_frame = PageArticle(
                    master=article_window,
                    db_conn=self.db_conn,
                    session_data=self.session_data
                )
            except TypeError:
                try:
                    page_frame = PageArticle(
                        article_window,
                        db_conn=self.db_conn,
                        session_data=self.session_data
                    )
                except TypeError:
                    page_frame = PageArticle(article_window)
            
            page_frame.grid(row=0, column=0, sticky="nsew")
            
            # Centrer
            article_window.update_idletasks()
            x = (article_window.winfo_screenwidth() // 2) - (600 // 2)
            y = (article_window.winfo_screenheight() // 2) - (700 // 2)
            article_window.geometry(f"600x700+{x}+{y}")
            
            article_window.focus()
            article_window.lift()
            
            # Attendre la fermeture et recharger les données
            article_window.wait_window()
            self.load_data()  # Recharger après fermeture
            
        except ImportError as e:
            messagebox.showerror(
                "Erreur d'import", 
                f"Impossible d'importer PageArticle.\n\nErreur: {e}"
            )
        except Exception as e:
            messagebox.showerror(
                "Erreur", 
                f"Impossible d'ouvrir la page article : {str(e)}"
            )

    def open_new_category(self):
        """Ouvre la page de création d'une nouvelle catégorie"""
        try:
            from pages.page_categorieArticle import PageCategorieArticle
            
            category_window = customtkinter.CTkToplevel(self)
            category_window.title("Nouvelle Catégorie")
            category_window.geometry("400x400")
            
            category_window.transient(self.master)
            category_window.grab_set()
            
            category_window.grid_columnconfigure(0, weight=1)
            category_window.grid_rowconfigure(0, weight=1)
            
            # Gestion des arguments pour la compatibilité
            try:
                page_frame = PageCategorieArticle(
                    master=category_window,
                    db_conn=self.db_conn,
                    session_data=self.session_data
                )
            except TypeError:
                try:
                    page_frame = PageCategorieArticle(
                        category_window,
                        db_conn=self.db_conn,
                        session_data=self.session_data
                    )
                except TypeError:
                    page_frame = PageCategorieArticle(category_window)
            
            page_frame.grid(row=0, column=0, sticky="nsew")
            
            # Centrer
            category_window.update_idletasks()
            x = (category_window.winfo_screenwidth() // 2) - (400 // 2)
            y = (category_window.winfo_screenheight() // 2) - (400 // 2)
            category_window.geometry(f"400x400+{x}+{y}")
            
            category_window.focus()
            category_window.lift()
            
            # Attendre la fermeture et recharger
            category_window.wait_window()
            self.load_data()
            
        except ImportError as e:
            messagebox.showerror(
                "Erreur d'import", 
                f"Impossible d'importer PageCategorieArticle.\n\nErreur: {e}"
            )
        except Exception as e:
            messagebox.showerror(
                "Erreur", 
                f"Impossible d'ouvrir la page catégorie : {str(e)}"
            )

    def load_data(self):
        """Récupère les données depuis la DB et les insère dans le Treeview."""
        # Effacer les données existantes
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        # Récupérer les données
        self.all_data = self.fetch_articles_from_db()
        
        # Insérer les données
        if self.all_data:
            for row in self.all_data:
                # S'assurer que les valeurs numériques sont formatées correctement si nécessaire
                # (Ici on suppose que psycopg2 retourne des chaînes ou nombres compatibles)
                self.tree.insert('', 'end', values=row)
            self.update_count(len(self.all_data))
        else:
            # S'assurer que le message d'absence de données est visible
            self.tree.insert('', 'end', values=("", "", "Aucun article trouvé", "", "", "", ""))
            self.update_count(0)

    def filter_data(self):
        """Filtre les données selon le critère de recherche"""
        # Effacer le Treeview
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Récupérer la valeur de recherche
        search_term = self.entry_search.get().lower().strip()
        
        # Si vide, afficher tout
        if not search_term:
            self.load_data()
            return
        
        # Filtrer
        filtered_data = []
        for row in self.all_data:
            # Concaténation des colonnes à rechercher (Code, Désignation, Unité, Catégorie)
            # Les indices correspondent : [1]Code, [2]Désignation, [3]Unité, [6]Catégorie
            searchable_text = f"{row[1]} {row[2]} {row[3]} {row[6]}".lower()
            
            if search_term in searchable_text:
                filtered_data.append(row)
        
        # Insérer
        if filtered_data:
            for row in filtered_data:
                self.tree.insert('', 'end', values=row)
            self.update_count(len(filtered_data))
        else:
            self.tree.insert('', 'end', values=("", "", "Aucun résultat trouvé", "", "", "", ""))
            self.update_count(0)

    def reset_filters(self):
        """Réinitialise le filtre"""
        self.entry_search.delete(0, 'end')
        self.load_data()

    def update_count(self, count):
        """Met à jour le compteur"""
        self.label_count.configure(text=f"Nombre d'articles : {count}")

    def export_to_excel(self):
        """Export vers Excel"""
        items = self.tree.get_children()
        
        if not items:
            messagebox.showwarning("Aucune donnée", "Aucune donnée à exporter.")
            return
        
        first_item = self.tree.item(items[0])
        # Vérifie si la première ligne est le message "Aucun article/résultat trouvé"
        if not first_item['values'] or first_item['values'][2] in ["Aucun article trouvé", "Aucun résultat trouvé"]:
            messagebox.showwarning("Aucune donnée", "Aucune donnée valide à exporter.")
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"Articles_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if not filename:
            return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Articles"
            
            headers = ["Code Article", "Désignation Article", "Unité", "Quantité", "Poids", "Catégorie"]
            ws.append(headers)
            
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            # Mise en forme des en-têtes
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Écriture des données
            valid_items_count = 0
            for item in items:
                # Les valeurs commencent à l'indice 1 pour exclure l'ID de la DB (index 0)
                values = self.tree.item(item)['values']
                if values and values[1]: # Assurez-vous qu'il y a un code article
                    ws.append(values[1:])
                    valid_items_count += 1
            
            # Définition des largeurs de colonnes
            column_widths = [15, 40, 12, 12, 12, 20]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + i)].width = width
            
            # Ligne de total
            ws.append([])
            total_row = ws.max_row
            ws[f'A{total_row}'] = f"Total: {valid_items_count} articles"
            ws[f'A{total_row}'].font = Font(bold=True, size=11)
            
            wb.save(filename)
            
            messagebox.showinfo(
                "Export réussi", 
                f"Les données ont été exportées avec succès vers:\n{filename}"
            )
            
        except PermissionError:
            messagebox.showerror(
                "Erreur d'accès",
                "Le fichier est peut-être ouvert dans Excel. Veuillez le fermer et réessayer."
            )
        except Exception as e:
            messagebox.showerror(
                "Erreur d'export",
                f"Une erreur est survenue lors de l'export:\n{e}"
            )


# ====================================================================
# 3. Classe Principale - Pour test standalone uniquement
# ====================================================================

if __name__ == "__main__":
    class App(customtkinter.CTk):
        def __init__(self):
            super().__init__()

            self.title("Application de Gestion des Articles")
            self.geometry("1100x600")
            customtkinter.set_appearance_mode("System") # Garder "System" pour un meilleur test
            
            self.grid_rowconfigure(1, weight=1)
            self.grid_columnconfigure(0, weight=1)

            label_title = customtkinter.CTkLabel(
                self, 
                text="Liste Complète des Articles", 
                font=customtkinter.CTkFont(size=20, weight="bold")
            )
            label_title.grid(row=0, column=0, padx=20, pady=(20, 10), sticky="nw")
            
            self.article_page = page_listeArticle(
                master=self, 
                db_conn=None,
                session_data=None
            )
            self.article_page.grid(row=1, column=0, sticky='nsew', padx=20, pady=(0, 20))

    app = App()
    app.mainloop()