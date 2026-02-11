import customtkinter as ctk
from tkinter import ttk, messagebox
import psycopg2
import json
import os
from datetime import datetime

class PageStockLivraison(ctk.CTkFrame):
    def __init__(self, master, db_conn=None, session_data=None, iduser=None):
        super().__init__(master)
        
        # Gestion de l'ID utilisateur
        if iduser is not None:
            self.iduser = iduser
        elif session_data and 'user_id' in session_data:
            self.iduser = session_data['user_id']
        else:
            self.iduser = 1
        
        self.magasins = []
        self.setup_ui()
        self.charger_magasins()
        self.charger_donnees()
    
    def connect_db(self):
        """Connexion qui remonte d'un niveau pour trouver config.json à la racine"""
        try:
            current_dir = os.path.dirname(os.path.abspath(__file__))
            root_dir = os.path.dirname(current_dir)
            config_path = os.path.join(root_dir, 'config.json')
 
            with open(config_path, 'r') as f:
                config = json.load(f)
                db_config = config['database']
            return psycopg2.connect(**db_config)
        except Exception as e:
            messagebox.showerror("Erreur de chemin", f"Impossible de trouver config.json à la racine.\nErreur: {e}")
            return None
    
    def formater_nombre(self, nombre):
        """Formate les nombres pour l'affichage (ex: 1.250,00)"""
        try:
            return f"{float(nombre):,.2f}".replace(',', ' ').replace('.', ',').replace(' ', '.')
        except:
            return "0,00"
    
    def setup_ui(self):
        """Création de l'interface utilisateur"""
        # Titre
        titre = ctk.CTkLabel(
            self, 
            text="📊 Stock et Livraisons en Attente", 
            font=ctk.CTkFont(size=20, weight="bold")
        )
        titre.pack(pady=10)
        
        # Frame de recherche et filtres
        frame_filtres = ctk.CTkFrame(self)
        frame_filtres.pack(fill="x", padx=20, pady=10)
        
        # Champ de recherche
        ctk.CTkLabel(frame_filtres, text="🔍 Recherche:", font=ctk.CTkFont(size=12)).pack(side="left", padx=5)
        self.entry_recherche = ctk.CTkEntry(
            frame_filtres, 
            placeholder_text="Code article ou Désignation...", 
            width=300
        )
        self.entry_recherche.pack(side="left", padx=5)
        self.entry_recherche.bind('<KeyRelease>', self.filtrer_donnees)
        
        # Combobox Magasin
        ctk.CTkLabel(frame_filtres, text="🏪 Magasin:", font=ctk.CTkFont(size=12)).pack(side="left", padx=(20, 5))
        self.combo_magasin = ctk.CTkComboBox(
            frame_filtres,
            values=["Tous"],
            width=200,
            command=self.filtrer_donnees
        )
        self.combo_magasin.pack(side="left", padx=5)
        self.combo_magasin.set("Tous")
        
        # Bouton Export Excel
        ctk.CTkButton(
            frame_filtres,
            text="📊 Export Excel",
            command=self.exporter_excel,
            fg_color="#0288d1",
            width=120
        ).pack(side="right", padx=5)
        
        # Bouton Actualiser
        ctk.CTkButton(
            frame_filtres,
            text="🔄 Actualiser",
            command=self.charger_donnees,
            fg_color="#2e7d32",
            width=120
        ).pack(side="right", padx=5)
        
        # Frame pour le tableau
        frame_tableau = ctk.CTkFrame(self)
        frame_tableau.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Création du Treeview
        colonnes = (
            "Code Article",
            "Désignation Article",
            "Unité",
            "Magasin",
            "Nom Client",
            "Reste à Livrer",
            "Stock Théorique",
            "Stock Réel"
        )
        
        self.tree = ttk.Treeview(
            frame_tableau,
            columns=colonnes,
            show="headings",
            height=20
        )
        
        # Configuration des colonnes
        largeurs = {
            "Code Article": 120,
            "Désignation Article": 250,
            "Unité": 100,
            "Magasin": 150,
            "Nom Client": 200,
            "Reste à Livrer": 120,
            "Stock Théorique": 130,
            "Stock Réel": 120
        }
        
        for col in colonnes:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=largeurs.get(col, 120), anchor='center')
        
        # Scrollbars
        scrollbar_y = ctk.CTkScrollbar(
            frame_tableau,
            orientation="vertical",
            command=self.tree.yview
        )
        scrollbar_x = ctk.CTkScrollbar(
            frame_tableau,
            orientation="horizontal",
            command=self.tree.xview
        )
        
        self.tree.configure(
            yscrollcommand=scrollbar_y.set,
            xscrollcommand=scrollbar_x.set
        )
        
        # Placement du tableau et scrollbars
        self.tree.grid(row=0, column=0, sticky="nsew")
        scrollbar_y.grid(row=0, column=1, sticky="ns")
        scrollbar_x.grid(row=1, column=0, sticky="ew")
        
        frame_tableau.grid_rowconfigure(0, weight=1)
        frame_tableau.grid_columnconfigure(0, weight=1)
        
        # Barre d'état
        frame_info = ctk.CTkFrame(self)
        frame_info.pack(fill="x", padx=20, pady=10)
        
        self.label_total = ctk.CTkLabel(
            frame_info,
            text="Total lignes: 0",
            font=ctk.CTkFont(size=12, weight="bold")
        )
        self.label_total.pack(side="left", padx=20)
        
        self.label_maj = ctk.CTkLabel(
            frame_info,
            text="Dernière MAJ: --",
            font=ctk.CTkFont(size=12)
        )
        self.label_maj.pack(side="right", padx=20)
    
    def charger_magasins(self):
        """Charge la liste des magasins pour le combobox"""
        conn = self.connect_db()
        if not conn:
            return
        
        try:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT idmag, designationmag 
                FROM tb_magasin 
                WHERE deleted = 0 
                ORDER BY designationmag
            """)
            self.magasins = cursor.fetchall()
            
            # Mise à jour du combobox
            valeurs_combo = ["Tous"] + [mag[1] for mag in self.magasins]
            self.combo_magasin.configure(values=valeurs_combo)
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur chargement magasins: {e}")
        finally:
            cursor.close()
            conn.close()
    
    def calculer_stock_theorique(self, idarticle, idunite, idmag):
        """
        Calcule le stock théorique d'un article pour un magasin
        (Réutilise la logique de page_stock.py)
        """
        conn = self.connect_db()
        if not conn:
            return 0
        
        try:
            cursor = conn.cursor()
            
            # Récupérer les unités de l'article
            cursor.execute("""
                SELECT idunite, COALESCE(qtunite, 1) as qtunite
                FROM tb_unite 
                WHERE idarticle = %s
                ORDER BY idunite ASC
            """, (idarticle,))
            unites_article = cursor.fetchall()
            
            if not unites_article:
                return 0
            
            idunite_base = unites_article[0][0]
            facteurs_conversion = {idunite_base: 1.0}
            
            facteur_cumul = 1.0
            for i, (id_unite, qt_unite) in enumerate(unites_article):
                if i == 0:
                    facteurs_conversion[id_unite] = 1.0
                else:
                    facteur_cumul *= qt_unite
                    facteurs_conversion[id_unite] = facteur_cumul
            
            stock_en_unite_base = 0
            
            # Calcul pour chaque unité
            for idunite_source, qtunite_source in unites_article:
                # Livraisons fournisseurs
                cursor.execute("""
                    SELECT COALESCE(SUM(qtlivrefrs), 0) 
                    FROM tb_livraisonfrs 
                    WHERE idarticle = %s AND idunite = %s AND idmag = %s
                """, (idarticle, idunite_source, idmag))
                total_livraison = cursor.fetchone()[0] or 0
                
                # Ventes
                cursor.execute("""
                    SELECT COALESCE(SUM(qtvente), 0) 
                    FROM tb_ventedetail 
                    WHERE idarticle = %s AND idunite = %s AND idmag = %s
                """, (idarticle, idunite_source, idmag))
                total_vente = cursor.fetchone()[0] or 0
                
                # Sorties
                cursor.execute("""
                    SELECT COALESCE(SUM(qtsortie), 0) 
                    FROM tb_sortiedetail 
                    WHERE idarticle = %s AND idunite = %s AND idmag = %s
                """, (idarticle, idunite_source, idmag))
                total_sortie = cursor.fetchone()[0] or 0
                
                # Transferts sortants
                cursor.execute("""
                    SELECT COALESCE(SUM(td.qttransfertsortie), 0)
                    FROM tb_transfertdetail td
                    INNER JOIN tb_transfert t ON td.idtransfert = t.idtransfert
                    WHERE td.idarticle = %s AND td.idunite = %s AND t.idmagsortie = %s
                """, (idarticle, idunite_source, idmag))
                total_transfert_sortie = cursor.fetchone()[0] or 0
                
                # Transferts entrants
                cursor.execute("""
                    SELECT COALESCE(SUM(td.qttransfertentree), 0)
                    FROM tb_transfertdetail td
                    INNER JOIN tb_transfert t ON td.idtransfert = t.idtransfert
                    WHERE td.idarticle = %s AND td.idunite = %s AND t.idmagentree = %s
                """, (idarticle, idunite_source, idmag))
                total_transfert_entree = cursor.fetchone()[0] or 0
                
                # Avoir
                cursor.execute("""
                    SELECT COALESCE(SUM(qtavoir), 0) 
                    FROM tb_avoirdetail 
                    WHERE idarticle = %s AND idunite = %s AND idmag = %s
                """, (idarticle, idunite_source, idmag))
                total_avoir = cursor.fetchone()[0] or 0
                
                stock_unite_source = (total_livraison + total_avoir + total_transfert_entree - 
                                     total_vente - total_sortie - total_transfert_sortie)
                
                facteur_vers_base = facteurs_conversion.get(idunite_source, 1.0)
                stock_en_unite_base += stock_unite_source * facteur_vers_base
            
            # Conversion vers l'unité cible
            facteur_cible = facteurs_conversion.get(idunite, 1.0)
            if facteur_cible == 0:
                return 0
            
            return stock_en_unite_base / facteur_cible
            
        except Exception as e:
            print(f"Erreur calcul stock théorique: {e}")
            return 0
        finally:
            cursor.close()
            conn.close()
    
    def charger_donnees(self):
        """Charge les données du tableau"""
        # Vider le tableau
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        conn = self.connect_db()
        if not conn:
            return
        
        try:
            cursor = conn.cursor()
            
            # Requête pour récupérer les livraisons clients en attente
            query = """
                SELECT 
                    u.codearticle,
                    a.designation,
                    u.designationunite,
                    m.designationmag,
                    c.nomcli,
                    l.idarticle,
                    l.idunite,
                    l.idmag,
                    COALESCE(SUM(l.qtvente - COALESCE(l.qtlivrecli, 0)), 0) as reste_a_livrer
                FROM tb_livraisoncli l
                INNER JOIN tb_article a ON l.idarticle = a.idarticle
                INNER JOIN tb_unite u ON (l.idarticle = u.idarticle AND l.idunite = u.idunite)
                INNER JOIN tb_magasin m ON l.idmag = m.idmag
                INNER JOIN tb_client c ON l.idclient = c.idclient
                WHERE a.deleted = 0 
                  AND (l.qtvente - COALESCE(l.qtlivrecli, 0)) > 0
                GROUP BY u.codearticle, a.designation, u.designationunite, m.designationmag, c.nomcli, 
                         l.idarticle, l.idunite, l.idmag
                ORDER BY u.codearticle, m.designationmag
            """
            
            cursor.execute(query)
            resultats = cursor.fetchall()
            
            for row in resultats:
                code_art = row[0]
                designation = row[1]
                unite = row[2]
                magasin = row[3]
                nom_client = row[4]
                idarticle = row[5]
                idunite = row[6]
                idmag = row[7]
                reste_a_livrer = row[8]
                
                # Calculer le stock théorique
                stock_theorique = self.calculer_stock_theorique(idarticle, idunite, idmag)
                
                # Calculer le stock réel
                stock_reel = stock_theorique + reste_a_livrer
                
                # Insertion dans le tableau
                values = (
                    code_art,
                    designation,
                    unite,
                    magasin,
                    nom_client,
                    self.formater_nombre(reste_a_livrer),
                    self.formater_nombre(stock_theorique),
                    self.formater_nombre(stock_reel)
                )
                
                self.tree.insert("", "end", values=values)
            
            # Mise à jour des labels
            self.label_total.configure(text=f"Total lignes: {len(resultats)}")
            self.label_maj.configure(text=f"Dernière MAJ: {datetime.now().strftime('%H:%M:%S')}")
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur chargement données: {e}")
            print(f"Détails erreur: {e}")
        finally:
            cursor.close()
            conn.close()
    
    def filtrer_donnees(self, event=None):
        """Filtre les données selon la recherche et le magasin sélectionné"""
        terme_recherche = self.entry_recherche.get().lower()
        magasin_filtre = self.combo_magasin.get()
        
        for item in self.tree.get_children():
            values = self.tree.item(item)['values']
            
            # Filtrage par recherche
            code = str(values[0]).lower()
            designation = str(values[1]).lower()
            correspond_recherche = (terme_recherche in code or terme_recherche in designation) if terme_recherche else True
            
            # Filtrage par magasin
            magasin = str(values[3])
            correspond_magasin = (magasin_filtre == "Tous" or magasin == magasin_filtre)
            
            # Afficher/masquer l'item
            if correspond_recherche and correspond_magasin:
                self.tree.reattach(item, '', 'end')
            else:
                self.tree.detach(item)
    
    def exporter_excel(self):
        """Exporte les données visibles vers un fichier Excel"""
        try:
            from tkinter import filedialog
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Vérifier s'il y a des données
            items_visibles = [item for item in self.tree.get_children() if self.tree.parent(item) == '']
            
            if not items_visibles:
                messagebox.showwarning("Attention", "Aucune donnée à exporter")
                return
            
            # Demander où enregistrer
            fichier = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"stock_livraisons_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if not fichier:
                return
            
            # Créer le classeur Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Stock et Livraisons"
            
            # Styles
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # En-têtes
            colonnes = [
                "Code Article",
                "Désignation Article",
                "Unité",
                "Magasin",
                "Nom Client",
                "Reste à Livrer",
                "Stock Théorique",
                "Stock Réel"
            ]
            
            for col_num, colonne in enumerate(colonnes, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = colonne
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            
            # Données
            for row_num, item in enumerate(items_visibles, 2):
                values = self.tree.item(item)['values']
                
                for col_num, value in enumerate(values, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Ajuster la largeur des colonnes
            largeurs = {
                1: 15,  # Code Article
                2: 35,  # Désignation
                3: 12,  # Unité
                4: 20,  # Magasin
                5: 25,  # Nom Client
                6: 18,  # Reste à Livrer
                7: 18,  # Stock Théorique
                8: 15   # Stock Réel
            }
            
            for col_num, largeur in largeurs.items():
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = largeur
            
            # Ajouter des informations supplémentaires
            derniere_ligne = len(items_visibles) + 3
            ws.cell(row=derniere_ligne, column=1).value = f"Exporté le: {datetime.now().strftime('%d/%m/%Y à %H:%M:%S')}"
            ws.cell(row=derniere_ligne, column=1).font = Font(italic=True, size=9)
            
            ws.cell(row=derniere_ligne + 1, column=1).value = f"Total lignes: {len(items_visibles)}"
            ws.cell(row=derniere_ligne + 1, column=1).font = Font(bold=True, size=10)
            
            # Enregistrer
            wb.save(fichier)
            
            messagebox.showinfo(
                "Succès",
                f"Export réussi !\n\n{len(items_visibles)} lignes exportées vers:\n{fichier}"
            )
            
        except ImportError:
            messagebox.showerror(
                "Erreur",
                "Le module 'openpyxl' n'est pas installé.\n\nInstallez-le avec:\npip install openpyxl"
            )
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'export Excel:\n{str(e)}")


# Test de la fenêtre
if __name__ == "__main__":
    app = ctk.CTk()
    app.geometry("1200x700")
    app.title("Gestion Stock et Livraisons")
    
    page = PageStockLivraison(app, iduser=1)
    page.pack(fill="both", expand=True)
    
    app.mainloop()