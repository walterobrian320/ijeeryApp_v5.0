# -*- coding: utf-8 -*-
"""
Page Voyage — iJeery V5.0
Suivi du chargement / déchargement des marchandises d'un camion ou d'un bateau.
Construite sur le même modèle que page_CmdFrs.py (thème app_theme, recherche
article/fournisseur identique, export Excel du bon de voyage).
"""

import os
import sys
import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog
import psycopg2
from datetime import datetime
from tkcalendar import DateEntry

# Permet de lancer ce fichier isolément depuis pages/ (ex: bouton "Run" de
# l'éditeur) en ajoutant la racine du projet au sys.path, comme le fait
# app_main.py avant de charger les pages en temps normal.
_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _BASE not in sys.path:
    sys.path.insert(0, _BASE)

from app_theme import Colors, Fonts, styled, Theme
from pages.ui_dialogs import YesNoDialog


# Correspondance affichage <-> valeur stockée en base (sans accent, encodage WIN1252)
STATUT_DB_TO_LABEL = {"Encours": "Encours", "Cloture": "Clôturé"}
STATUT_LABEL_TO_DB = {"Encours": "Encours", "Clôturé": "Cloture"}


# ─────────────────────────────────────────────────────────────────────────────
def _apply_treeview_style():
    style = ttk.Style()
    style.theme_use("clam")
    style.configure(
        "iJeery.Treeview",
        background=Colors.BG_CARD,
        foreground=Colors.TEXT_PRIMARY,
        fieldbackground=Colors.BG_CARD,
        rowheight=28,
        font=("Segoe UI", 9),
        borderwidth=0,
    )
    style.configure(
        "iJeery.Treeview.Heading",
        background=Colors.MIDNIGHT,
        foreground=Colors.TEXT_ON_DARK,
        font=("Segoe UI", 9, "bold"),
        relief="flat",
        borderwidth=0,
    )
    style.map("iJeery.Treeview",
              background=[("selected", Colors.PRIMARY_LIGHT)],
              foreground=[("selected", Colors.TEXT_PRIMARY)])
    style.map("iJeery.Treeview.Heading",
              background=[("active", Colors.MIDNIGHT_LIGHT)])


class PageVoyage(ctk.CTkFrame):
    def __init__(self, parent, iduser):
        super().__init__(parent, fg_color=Colors.BG_PAGE)
        _apply_treeview_style()
        self.iduser = iduser

        self.vehicules = {}          # id_vehicule -> {marque, immatriculation, type_vehicule}
        self.vehicule_id = None
        self.fournisseur_id = None
        self.article_selectionne = None
        self.items_voyage = []

        self.idvoyage_charge = None
        self.mode_modification = False
        self.index_ligne_selectionnee = None

        self.setup_ui()
        self.charger_vehicules()
        self.generer_numero_voyage()

    # ─────────────────────────────────────────────────────────────────────
    # DB
    # ─────────────────────────────────────────────────────────────────────
    def connect_db(self):
        try:
            from pages.db_helper import connect_page_db
            return connect_page_db()
        except FileNotFoundError:
            messagebox.showerror("Erreur", "Fichier 'config.json' non trouvé.")
        except KeyError:
            messagebox.showerror("Erreur", "Clés DB manquantes dans 'config.json'.")
        except psycopg2.Error as e:
            messagebox.showerror("Connexion", f"Erreur PostgreSQL : {e}")
        except UnicodeDecodeError as e:
            messagebox.showerror("Encodage", f"Problème d'encodage : {e}")
        return None

    # ─────────────────────────────────────────────────────────────────────
    # Utilitaires numériques
    # ─────────────────────────────────────────────────────────────────────
    def formater_nombre(self, nombre):
        try:
            nombre = float(nombre)
            partie_entiere = int(nombre)
            partie_decimale = abs(nombre - partie_entiere)
            str_entiere = f"{partie_entiere:,}".replace(',', '.')
            str_decimale = f"{partie_decimale:.2f}".split('.')[1]
            return f"{str_entiere},{str_decimale}"
        except Exception:
            return "0,00"

    def parser_nombre(self, texte):
        try:
            return float(texte.replace('.', '').replace(',', '.'))
        except Exception:
            return 0.0

    def _format_date_db(self, date_str):
        """Convertit JJ/MM/AAAA → AAAA-MM-JJ pour la base de données."""
        if not date_str:
            return None
        try:
            if '/' in date_str:
                p = date_str.split('/')
                if len(p) == 3:
                    return f"{p[2]}-{p[1]}-{p[0]}"
            return date_str
        except Exception:
            return None

    def _configure_table_alternating_colors(self, tree):
        tree.tag_configure("row_even", background=Colors.BG_CARD)
        tree.tag_configure("row_odd", background=Colors.BG_ROW_ALT)

    def _refresh_table_alternating_colors(self, tree):
        for idx, item in enumerate(tree.get_children()):
            tags = tuple(t for t in tree.item(item, "tags") if t not in ("row_even", "row_odd"))
            alt = "row_even" if idx % 2 == 0 else "row_odd"
            tree.item(item, tags=tags + (alt,))

    # =========================================================================
    # CONSTRUCTION DE L'UI
    # =========================================================================
    def setup_ui(self):
        header = ctk.CTkFrame(self, fg_color=Colors.MIDNIGHT, corner_radius=0, height=42)
        header.pack(fill="x")
        header.pack_propagate(False)

        styled.button_info(
            header, text="Charger", icon="📂",
            command=self.ouvrir_recherche_voyage, width=120, height=28
        ).pack(side="right", padx=8, pady=7)

        styled.button_secondary(
            header, text="Nouveau", icon="🔄",
            command=self.nouveau_voyage, width=100, height=28
        ).pack(side="right", padx=0, pady=7)

        body = ctk.CTkScrollableFrame(
            self, fg_color=Colors.BG_PAGE,
            scrollbar_button_color=Colors.SILVER,
            scrollbar_button_hover_color=Colors.PRIMARY,
        )
        body.pack(fill="both", expand=True, padx=8, pady=6)

        self._build_section_infos(body)
        self._build_section_article(body)
        self._build_section_tableau(body)

    # ─────────────────────────────────────────────────────────────────────
    # Section 1 — Informations générales du voyage
    # ─────────────────────────────────────────────────────────────────────
    def _build_section_infos(self, parent):
        card = ctk.CTkFrame(parent, fg_color=Colors.BG_CARD,
                            corner_radius=8, border_width=1, border_color=Colors.BORDER)
        card.pack(fill="x", pady=(0, 4))

        # Ligne 1 : Voyage N° | Matériel de transport | Type véhicule
        r1 = ctk.CTkFrame(card, fg_color="transparent")
        r1.pack(fill="x", padx=10, pady=(8, 4))

        ctk.CTkLabel(r1, text="🚚 Voyage", font=Fonts.bold(11),
                     text_color=Colors.MIDNIGHT, width=80, anchor="w").pack(side="left", padx=(0, 6))

        def _field_wrap(par, label):
            wrap = ctk.CTkFrame(par, fg_color="transparent")
            wrap.pack(side="left", padx=(0, 10))
            ctk.CTkLabel(wrap, text=label, font=Fonts.small(9),
                         text_color=Colors.TEXT_MUTED).pack(anchor="w")
            return wrap

        w = _field_wrap(r1, "Voyage N°")
        self.entry_numero = ctk.CTkEntry(w, width=150, height=28,
                                         fg_color=Colors.BG_INPUT, border_color=Colors.BORDER,
                                         font=Fonts.body(11), state="readonly")
        self.entry_numero.pack()

        w = _field_wrap(r1, "Matériel de transport")
        self.combo_vehicule = ctk.CTkComboBox(
            w, width=220, height=28, values=["—"],
            fg_color=Colors.BG_INPUT, border_color=Colors.BORDER,
            button_color=Colors.PRIMARY, button_hover_color=Colors.PRIMARY_HOVER,
            dropdown_fg_color=Colors.BG_CARD, font=Fonts.body(11),
            state="readonly", command=self.on_vehicule_change
        )
        self.combo_vehicule.pack()

        w = _field_wrap(r1, "Type véhicule")
        self.entry_type_vehicule = ctk.CTkEntry(w, width=140, height=28,
                                                fg_color=Colors.BG_INPUT, border_color=Colors.BORDER,
                                                font=Fonts.body(11), state="readonly")
        self.entry_type_vehicule.pack()

        # Ligne 2 : Date création | Date clôture | Statut
        r2 = ctk.CTkFrame(card, fg_color="transparent")
        r2.pack(fill="x", padx=10, pady=4)
        ctk.CTkLabel(r2, text="", width=80).pack(side="left", padx=(0, 6))

        w = _field_wrap(r2, "Date de création")
        self.date_creation = DateEntry(w, width=11, background=Colors.MIDNIGHT,
                                       foreground="white", borderwidth=1,
                                       date_pattern="dd/mm/yyyy", font=("Segoe UI", 9))
        self.date_creation.pack()

        w = _field_wrap(r2, "Date de clôture")
        self.date_cloture = DateEntry(w, width=11, background=Colors.MIDNIGHT,
                                      foreground="white", borderwidth=1,
                                      date_pattern="dd/mm/yyyy", font=("Segoe UI", 9))
        self.date_cloture.pack()

        w = _field_wrap(r2, "Statut")
        self.combo_statut = ctk.CTkComboBox(
            w, width=130, height=28, values=["Encours", "Clôturé"],
            fg_color=Colors.BG_INPUT, border_color=Colors.BORDER,
            button_color=Colors.PRIMARY, button_hover_color=Colors.PRIMARY_HOVER,
            dropdown_fg_color=Colors.BG_CARD, font=Fonts.body(11),
            state="readonly"
        )
        self.combo_statut.set("Encours")
        self.combo_statut.pack()

        # Ligne 3 : Itinéraire
        r3 = ctk.CTkFrame(card, fg_color="transparent")
        r3.pack(fill="x", padx=10, pady=(4, 8))
        ctk.CTkLabel(r3, text="Itinéraire :", font=Fonts.label(10),
                     text_color=Colors.TEXT_SECONDARY, width=80, anchor="w").pack(side="left", padx=(0, 6))
        self.entry_itineraire = ctk.CTkEntry(
            r3, height=28, fg_color=Colors.BG_INPUT, border_color=Colors.BORDER,
            font=Fonts.body(11), placeholder_text="Ex: Toamasina - Antananarivo"
        )
        self.entry_itineraire.pack(side="left", fill="x", expand=True)

    def on_vehicule_change(self, choice=None):
        info = self.vehicules.get(self.combo_vehicule.get())
        self.entry_type_vehicule.configure(state="normal")
        self.entry_type_vehicule.delete(0, "end")
        if info:
            self.vehicule_id = info["id"]
            self.entry_type_vehicule.insert(0, info["type_vehicule"] or "")
        else:
            self.vehicule_id = None
        self.entry_type_vehicule.configure(state="readonly")

    def charger_vehicules(self):
        conn = self.connect_db()
        if not conn:
            return
        try:
            cur = conn.cursor()
            cur.execute("""
                SELECT id, marque, immatriculation, type_vehicule
                FROM logistique_vehicule
                WHERE statut = 'Actif'
                ORDER BY marque
            """)
            rows = cur.fetchall()
            self.vehicules = {}
            valeurs = []
            for r in rows:
                cle = f"{r[1]} — {r[2]}"
                self.vehicules[cle] = {"id": r[0], "type_vehicule": r[3] or ""}
                valeurs.append(cle)
            if not valeurs:
                valeurs = ["—"]
            self.combo_vehicule.configure(values=valeurs)
            self.combo_vehicule.set(valeurs[0])
            self.on_vehicule_change()
        except Exception as e:
            messagebox.showerror("Erreur", f"Véhicules : {e}")
        finally:
            if "cur" in locals():
                cur.close()
            conn.close()

    # ─────────────────────────────────────────────────────────────────────
    # Section 2 — Saisie des marchandises
    # ─────────────────────────────────────────────────────────────────────
    def _build_section_article(self, parent):
        card = ctk.CTkFrame(parent, fg_color=Colors.BG_CARD,
                            corner_radius=8, border_width=1, border_color=Colors.BORDER)
        card.pack(fill="x", pady=(0, 4))

        r1 = ctk.CTkFrame(card, fg_color="transparent")
        r1.pack(fill="x", padx=10, pady=(6, 2))
        r1.columnconfigure(2, weight=1)

        ctk.CTkLabel(r1, text="📦 Marchandise", font=Fonts.bold(11),
                     text_color=Colors.MIDNIGHT, width=100, anchor="w"
                     ).grid(row=0, column=0, sticky="w", padx=(0, 6))

        ctk.CTkLabel(r1, text="Provenance :", font=Fonts.label(10),
                     text_color=Colors.TEXT_SECONDARY).grid(row=0, column=1, sticky="w", padx=(0, 4))

        self.entry_provenance = ctk.CTkEntry(r1, height=28,
                                             fg_color=Colors.BG_INPUT, border_color=Colors.BORDER,
                                             font=Fonts.body(11), state="readonly")
        self.entry_provenance.grid(row=0, column=2, sticky="ew", padx=(0, 4))

        ctk.CTkButton(r1, text="🔍 Rechercher", width=110, height=28,
                      fg_color=Colors.PRIMARY, hover_color=Colors.PRIMARY_HOVER,
                      text_color="white", font=Fonts.body(10), corner_radius=6,
                      command=self.ouvrir_recherche_fournisseur
                      ).grid(row=0, column=3, sticky="w")

        r1b = ctk.CTkFrame(card, fg_color="transparent")
        r1b.pack(fill="x", padx=10, pady=(2, 2))
        r1b.columnconfigure(2, weight=1)
        ctk.CTkLabel(r1b, text="Désignation :", font=Fonts.label(10),
                     text_color=Colors.TEXT_SECONDARY, width=100, anchor="w"
                     ).grid(row=0, column=0, sticky="w", padx=(0, 6))

        self.entry_article = ctk.CTkEntry(r1b, height=28,
                                          fg_color=Colors.BG_INPUT, border_color=Colors.BORDER,
                                          font=Fonts.body(11), state="readonly")
        self.entry_article.grid(row=0, column=1, columnspan=2, sticky="ew", padx=(0, 4))

        ctk.CTkButton(r1b, text="🔍 Rechercher", width=110, height=28,
                      fg_color=Colors.PRIMARY, hover_color=Colors.PRIMARY_HOVER,
                      text_color="white", font=Fonts.body(10), corner_radius=6,
                      command=self.ouvrir_recherche_article
                      ).grid(row=0, column=3, sticky="w")

        # Ligne quantités / poids / prix
        r2 = ctk.CTkFrame(card, fg_color="transparent")
        r2.pack(fill="x", padx=10, pady=(2, 6))

        def _field(par, label, width=90):
            wrap = ctk.CTkFrame(par, fg_color="transparent")
            wrap.pack(side="left", padx=(0, 6))
            ctk.CTkLabel(wrap, text=label, font=Fonts.small(9),
                         text_color=Colors.TEXT_MUTED).pack(anchor="w")
            e = ctk.CTkEntry(wrap, width=width, height=26,
                             fg_color=Colors.BG_INPUT, border_color=Colors.BORDER,
                             font=Fonts.body(10))
            e.pack()
            return e

        self.entry_quantite = _field(r2, "Quantité", 90)
        self.entry_poids_unitaire = _field(r2, "Poids Unitaire", 100)
        self.entry_prix_unitaire = _field(r2, "Prix Unitaire", 100)
        self.entry_prix_vente = _field(r2, "Prix de Vente", 100)

        self.entry_quantite.bind('<KeyRelease>', lambda e: self.calculer_poids_total_preview())
        self.entry_poids_unitaire.bind('<KeyRelease>', lambda e: self.calculer_poids_total_preview())

        tot_wrap = ctk.CTkFrame(r2, fg_color="transparent")
        tot_wrap.pack(side="left", padx=(6, 10))
        ctk.CTkLabel(tot_wrap, text="Poids Total", font=Fonts.small(9),
                     text_color=Colors.TEXT_MUTED).pack(anchor="w")
        self.label_poids_total_ligne = ctk.CTkLabel(
            tot_wrap, text="0,00",
            font=Fonts.bold(10), text_color=Colors.SUCCESS_TEXT,
            fg_color=Colors.SUCCESS_LIGHT, corner_radius=5, padx=7, pady=3
        )
        self.label_poids_total_ligne.pack(anchor="w")

        btn_wrap = ctk.CTkFrame(r2, fg_color="transparent")
        btn_wrap.pack(side="right")
        ctk.CTkLabel(btn_wrap, text=" ", font=Fonts.small(9)).pack(anchor="w")
        btn_inner = ctk.CTkFrame(btn_wrap, fg_color="transparent")
        btn_inner.pack()

        self.btn_ajouter = ctk.CTkButton(
            btn_inner, text="➕ Ajouter", width=90, height=26,
            fg_color=Colors.SUCCESS, hover_color=Colors.SUCCESS_DARK,
            text_color=Colors.TEXT_ON_DARK, font=Fonts.button(10),
            corner_radius=5, command=self.ajouter_article
        )
        self.btn_ajouter.pack(side="left", padx=(0, 4))

        self.btn_modifier_ligne = ctk.CTkButton(
            btn_inner, text="✏️ Modif.", width=85, height=26,
            fg_color=Colors.WARNING, hover_color="#D68910",
            text_color=Colors.TEXT_ON_DARK, font=Fonts.button(10),
            corner_radius=5, state="disabled",
            command=self.modifier_ligne_article
        )
        self.btn_modifier_ligne.pack(side="left", padx=(0, 4))

        self.btn_annuler_selection = ctk.CTkButton(
            btn_inner, text="✖", width=28, height=26,
            fg_color=Colors.CLOUDS, hover_color=Colors.SILVER,
            text_color=Colors.TEXT_PRIMARY, font=Fonts.button(10),
            border_width=1, border_color=Colors.BORDER, corner_radius=5,
            state="disabled", command=self.annuler_selection_ligne
        )
        self.btn_annuler_selection.pack(side="left")

        # Ligne 3 : Péremption (optionnelle) / Destination / N° Facture / N° Camion / Chauffeur
        r3 = ctk.CTkFrame(card, fg_color="transparent")
        r3.pack(fill="x", padx=10, pady=(0, 6))

        per_wrap = ctk.CTkFrame(r3, fg_color="transparent")
        per_wrap.pack(side="left", padx=(0, 6))
        per_label_row = ctk.CTkFrame(per_wrap, fg_color="transparent")
        per_label_row.pack(anchor="w")

        self.var_has_peremption = ctk.BooleanVar(value=False)
        self.check_peremption = ctk.CTkCheckBox(
            per_label_row, text="", variable=self.var_has_peremption,
            command=self.toggle_date_peremption,
            checkbox_width=16, checkbox_height=16,
            checkmark_color=Colors.TEXT_ON_DARK,
            fg_color=Colors.PRIMARY, hover_color=Colors.PRIMARY_HOVER,
            width=18
        )
        self.check_peremption.pack(side="left", padx=(0, 2))
        ctk.CTkLabel(per_label_row, text="Péremption", font=Fonts.small(9),
                     text_color=Colors.TEXT_MUTED).pack(side="left")
        self.entry_peremption = DateEntry(
            per_wrap, width=9, background=Colors.MIDNIGHT,
            foreground="white", borderwidth=1,
            date_pattern="dd/mm/yyyy", state="disabled",
            font=("Segoe UI", 9)
        )
        self.entry_peremption.pack()

        self.entry_destination = _field(r3, "Destination", 130)
        self.entry_num_facture = _field(r3, "N° Facture", 110)
        self.entry_num_camion = _field(r3, "N° Camion", 100)
        self.entry_chauffeur = _field(r3, "Chauffeur", 130)

    def toggle_date_peremption(self, etat=None):
        if etat is not None:
            self.var_has_peremption.set(etat)
        if self.var_has_peremption.get():
            self.entry_peremption.configure(state="normal")
        else:
            self.entry_peremption.configure(state="disabled")

    def ouvrir_recherche_fournisseur(self):
        """Recherche de la provenance (fournisseur) — même architecture que la recherche d'article."""
        if self.index_ligne_selectionnee is not None:
            messagebox.showwarning("Attention", "Validez ou annulez la modification en cours.")
            return
        fen = ctk.CTkToplevel(self)
        fen.title("Rechercher un fournisseur")
        fen.geometry("800x420")
        fen.grab_set()
        Theme.apply_toplevel(fen)

        main = ctk.CTkFrame(fen, fg_color=Colors.BG_PAGE)
        main.pack(fill="both", expand=True, padx=12, pady=12)

        ctk.CTkLabel(main, text="Sélectionner un fournisseur (provenance)",
                     font=Fonts.heading(14), text_color=Colors.MIDNIGHT).pack(pady=(0, 10))

        sf = ctk.CTkFrame(main, fg_color="transparent")
        sf.pack(fill="x", pady=(0, 8))
        ctk.CTkLabel(sf, text="🔍").pack(side="left", padx=6)
        entry_s = ctk.CTkEntry(sf, placeholder_text="Nom ou contact...", height=34,
                               fg_color=Colors.BG_INPUT, border_color=Colors.BORDER, font=Fonts.body(11))
        entry_s.pack(side="left", fill="x", expand=True, padx=4)

        tf = ctk.CTkFrame(main, fg_color=Colors.BORDER, corner_radius=8)
        tf.pack(fill="both", expand=True, pady=(0, 8))
        cols = ("ID", "Nom", "Contact", "Adresse")
        tree = ttk.Treeview(tf, columns=cols, show="headings", height=10, style="iJeery.Treeview")
        self._configure_table_alternating_colors(tree)
        tree.column("ID", width=0, stretch=False)
        tree.column("Nom", width=180)
        tree.column("Contact", width=150)
        tree.column("Adresse", width=280)
        for c in cols:
            tree.heading(c, text=c)
        sb = ttk.Scrollbar(tf, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=sb.set)
        tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        lbl_c = ctk.CTkLabel(main, text="", font=Fonts.small(10), text_color=Colors.TEXT_MUTED)
        lbl_c.pack(pady=(0, 4))

        def charger(filtre=""):
            for i in tree.get_children():
                tree.delete(i)
            conn = self.connect_db()
            if not conn:
                return
            try:
                cur = conn.cursor()
                q = "SELECT idfrs, nomfrs, contactfrs, adressefrs FROM tb_fournisseur WHERE deleted=0"
                p = []
                if filtre:
                    q += " AND (LOWER(nomfrs) LIKE LOWER(%s) OR LOWER(contactfrs) LIKE LOWER(%s))"
                    p = [f"%{filtre}%", f"%{filtre}%"]
                q += " ORDER BY nomfrs"
                cur.execute(q, p)
                rows = cur.fetchall()
                for r in rows:
                    tree.insert("", "end", values=(r[0], r[1] or "", r[2] or "", r[3] or ""))
                self._refresh_table_alternating_colors(tree)
                lbl_c.configure(text=f"{len(rows)} fournisseur(s)")
            except Exception as e:
                messagebox.showerror("Erreur", str(e))
            finally:
                conn.close()

        entry_s.bind('<KeyRelease>', lambda e: charger(entry_s.get()))

        def valider():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Attention", "Sélectionnez un fournisseur.")
                return
            v = tree.item(sel[0])['values']
            self.fournisseur_id = v[0]
            self.entry_provenance.configure(state="normal")
            self.entry_provenance.delete(0, "end")
            self.entry_provenance.insert(0, v[1])
            self.entry_provenance.configure(state="readonly")
            fen.destroy()

        tree.bind('<Double-Button-1>', lambda e: valider())
        bf = ctk.CTkFrame(main, fg_color="transparent")
        bf.pack(fill="x")
        styled.button_danger(bf, text="Annuler", icon="❌", width=110, height=36, command=fen.destroy).pack(side="left", padx=4)
        styled.button_success(bf, text="Valider", icon="✅", width=110, height=36, command=valider).pack(side="right", padx=4)
        charger()

    def calculer_poids_total_preview(self):
        try:
            t = self.parser_nombre(self.entry_quantite.get()) * self.parser_nombre(self.entry_poids_unitaire.get())
            self.label_poids_total_ligne.configure(text=self.formater_nombre(t))
        except Exception:
            self.label_poids_total_ligne.configure(text="0,00")

    # ─────────────────────────────────────────────────────────────────────
    # Recherche d'article (identique au mécanisme de page_CmdFrs.py)
    # ─────────────────────────────────────────────────────────────────────
    def ouvrir_recherche_article(self):
        if self.index_ligne_selectionnee is not None:
            messagebox.showwarning("Attention", "Validez ou annulez la modification en cours.")
            return
        fen = ctk.CTkToplevel(self)
        fen.title("Rechercher un article")
        fen.geometry("1000x580")
        fen.grab_set()
        Theme.apply_toplevel(fen)

        main = ctk.CTkFrame(fen, fg_color=Colors.BG_PAGE)
        main.pack(fill="both", expand=True, padx=12, pady=12)

        ctk.CTkLabel(main, text="Sélectionner un article",
                     font=Fonts.heading(14), text_color=Colors.MIDNIGHT).pack(pady=(0, 10))

        sf = ctk.CTkFrame(main, fg_color="transparent")
        sf.pack(fill="x", pady=(0, 8))
        ctk.CTkLabel(sf, text="🔍").pack(side="left", padx=6)
        entry_s = ctk.CTkEntry(sf, placeholder_text="Code ou désignation...", height=34,
                               fg_color=Colors.BG_INPUT, border_color=Colors.BORDER, font=Fonts.body(11))
        entry_s.pack(side="left", fill="x", expand=True, padx=4)

        tf = ctk.CTkFrame(main, fg_color=Colors.BORDER, corner_radius=8)
        tf.pack(fill="both", expand=True, pady=(0, 8))
        cols = ("ID_Article", "ID_Unite", "Code", "Désignation", "Unité")
        tree = ttk.Treeview(tf, columns=cols, show="headings", height=14, style="iJeery.Treeview")
        self._configure_table_alternating_colors(tree)
        tree.column("ID_Article", width=0, stretch=False)
        tree.column("ID_Unite", width=0, stretch=False)
        tree.column("Code", width=140)
        tree.column("Désignation", width=480)
        tree.column("Unité", width=110)
        for c in cols:
            tree.heading(c, text=c)
        sb = ttk.Scrollbar(tf, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=sb.set)
        tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        lbl_c = ctk.CTkLabel(main, text="", font=Fonts.small(10), text_color=Colors.TEXT_MUTED)
        lbl_c.pack(pady=(0, 4))

        def charger(filtre=""):
            for i in tree.get_children():
                tree.delete(i)
            conn = self.connect_db()
            if not conn:
                return
            try:
                cur = conn.cursor()
                q = """SELECT T2.idarticle, T1.codearticle, T2.designation, T1.designationunite, T1.idunite
                       FROM tb_unite T1 INNER JOIN tb_article T2 ON T1.idarticle=T2.idarticle
                       WHERE T2.deleted=0"""
                p = []
                if filtre:
                    q += " AND (LOWER(T1.codearticle) LIKE LOWER(%s) OR LOWER(T2.designation) LIKE LOWER(%s))"
                    p = [f"%{filtre}%", f"%{filtre}%"]
                q += ' ORDER BY T1.codearticle'
                cur.execute(q, p)
                rows = cur.fetchall()
                for r in rows:
                    tree.insert("", "end", values=(r[0], r[4], r[1], r[2], r[3]))
                self._refresh_table_alternating_colors(tree)
                lbl_c.configure(text=f"{len(rows)} article(s)")
            except Exception as e:
                messagebox.showerror("Erreur", str(e))
            finally:
                conn.close()

        entry_s.bind('<KeyRelease>', lambda e: charger(entry_s.get()))

        def valider():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Attention", "Sélectionnez un article.")
                return
            v = tree.item(sel[0])['values']
            self.article_selectionne = {'idarticle': v[0], 'idunite': v[1], 'nomart': v[3], 'unite': v[4]}
            self.entry_article.configure(state="normal")
            self.entry_article.delete(0, "end")
            self.entry_article.insert(0, v[3])
            self.entry_article.configure(state="readonly")
            self.entry_quantite.delete(0, "end")
            self.entry_poids_unitaire.delete(0, "end")
            self.entry_prix_unitaire.delete(0, "end")
            self.entry_prix_vente.delete(0, "end")
            self.calculer_poids_total_preview()
            fen.destroy()

        tree.bind('<Double-Button-1>', lambda e: valider())
        bf = ctk.CTkFrame(main, fg_color="transparent")
        bf.pack(fill="x")
        styled.button_danger(bf, text="Annuler", icon="❌", width=110, height=36, command=fen.destroy).pack(side="left", padx=4)
        styled.button_success(bf, text="Valider", icon="✅", width=110, height=36, command=valider).pack(side="right", padx=4)
        charger()

    # ─────────────────────────────────────────────────────────────────────
    # Ajouter / Modifier / Supprimer une ligne
    # ─────────────────────────────────────────────────────────────────────
    def ajouter_article(self):
        if not self.article_selectionne:
            messagebox.showwarning("Attention", "Sélectionnez un article.")
            return
        if not self.fournisseur_id:
            messagebox.showwarning("Attention", "Sélectionnez une provenance.")
            return
        prov = self.entry_provenance.get()
        idfrs = self.fournisseur_id
        try:
            quantite = self.parser_nombre(self.entry_quantite.get())
            poids_unitaire = self.parser_nombre(self.entry_poids_unitaire.get())
            prix_unitaire = self.parser_nombre(self.entry_prix_unitaire.get())
            prix_vente = self.parser_nombre(self.entry_prix_vente.get())
            if quantite <= 0:
                messagebox.showwarning("Attention", "La quantité doit être > 0.")
                return
            poids_total = quantite * poids_unitaire
            date_p = self.entry_peremption.get_date().strftime('%d/%m/%Y') if self.var_has_peremption.get() else ""
            destination = self.entry_destination.get().strip()
            num_facture = self.entry_num_facture.get().strip()
            num_camion = self.entry_num_camion.get().strip()
            chauffeur = self.entry_chauffeur.get().strip()

            self.tree.insert("", "end", values=(
                prov, self.article_selectionne['nomart'], self.article_selectionne['unite'],
                self.formater_nombre(quantite), self.formater_nombre(poids_unitaire),
                self.formater_nombre(prix_unitaire), self.formater_nombre(poids_total),
                self.formater_nombre(prix_vente), date_p, destination, num_facture, num_camion, chauffeur
            ))
            self._refresh_table_alternating_colors(self.tree)

            self.items_voyage.append({
                'iddetail': None, 'idfrs': idfrs, 'nomfrs': prov,
                'idarticle': self.article_selectionne['idarticle'],
                'idunite': self.article_selectionne['idunite'],
                'nomart': self.article_selectionne['nomart'], 'unite': self.article_selectionne['unite'],
                'quantite': quantite, 'poids_unitaire': poids_unitaire,
                'prix_unitaire': prix_unitaire, 'poids_total': poids_total,
                'prix_vente': prix_vente, 'date_peremption': date_p or None,
                'destination': destination, 'num_facture': num_facture,
                'num_camion': num_camion, 'chauffeur': chauffeur
            })

            self._reset_champs_article()
            self.calculer_total()
        except ValueError:
            messagebox.showerror("Erreur", "Données numériques invalides.")

    def _reset_champs_article(self):
        for e in (self.entry_article, self.entry_provenance):
            e.configure(state="normal")
            e.delete(0, "end")
            e.configure(state="readonly")
        for e in (self.entry_quantite, self.entry_poids_unitaire, self.entry_prix_unitaire, self.entry_prix_vente,
                  self.entry_destination, self.entry_num_facture, self.entry_num_camion, self.entry_chauffeur):
            e.delete(0, "end")
        self.toggle_date_peremption(False)
        self.article_selectionne = None
        self.fournisseur_id = None
        self.label_poids_total_ligne.configure(text="0,00")

    def on_selection_ligne(self, event):
        if self.tree.selection():
            self.btn_modifier_ligne.configure(state="normal")
            self.btn_annuler_selection.configure(state="normal")

    def on_double_click_ligne(self, event):
        sel = self.tree.selection()
        if sel:
            self.charger_ligne_pour_modification(sel[0])

    def charger_ligne_pour_modification(self, item_id):
        index = self.tree.index(item_id)
        self.index_ligne_selectionnee = index
        item = self.items_voyage[index]

        self.fournisseur_id = item.get('idfrs')
        self.entry_provenance.configure(state="normal")
        self.entry_provenance.delete(0, "end")
        self.entry_provenance.insert(0, item.get('nomfrs') or "")
        self.entry_provenance.configure(state="readonly")

        self.entry_article.configure(state="normal")
        self.entry_article.delete(0, "end")
        self.entry_article.insert(0, item['nomart'])
        self.entry_article.configure(state="readonly")

        self.entry_quantite.delete(0, "end")
        self.entry_quantite.insert(0, self.formater_nombre(item['quantite']))
        self.entry_poids_unitaire.delete(0, "end")
        self.entry_poids_unitaire.insert(0, self.formater_nombre(item['poids_unitaire']))
        self.entry_prix_unitaire.delete(0, "end")
        self.entry_prix_unitaire.insert(0, self.formater_nombre(item['prix_unitaire']))
        self.entry_prix_vente.delete(0, "end")
        self.entry_prix_vente.insert(0, self.formater_nombre(item['prix_vente']))

        for e, cle in ((self.entry_destination, 'destination'), (self.entry_num_facture, 'num_facture'),
                       (self.entry_num_camion, 'num_camion'), (self.entry_chauffeur, 'chauffeur')):
            e.delete(0, "end")
            e.insert(0, item.get(cle) or "")

        if item.get('date_peremption'):
            self.toggle_date_peremption(True)
            try:
                p = item['date_peremption'].split('/')
                if len(p) == 3:
                    self.entry_peremption.set_date(datetime(int(p[2]), int(p[1]), int(p[0])))
            except Exception:
                pass
        else:
            self.toggle_date_peremption(False)

        self.article_selectionne = {
            'idarticle': item['idarticle'], 'idunite': item['idunite'],
            'nomart': item['nomart'], 'unite': item['unite']
        }

        self.btn_ajouter.configure(state="disabled")
        self.btn_modifier_ligne.configure(state="normal", text="✅  Valider Modif.")
        self.btn_annuler_selection.configure(state="normal")
        self.calculer_poids_total_preview()

    def modifier_ligne_article(self):
        sel = self.tree.selection()
        if self.index_ligne_selectionnee is None:
            if sel:
                self.charger_ligne_pour_modification(sel[0])
            else:
                messagebox.showwarning("Attention", "Sélectionnez une ligne à modifier.")
            return
        try:
            quantite = self.parser_nombre(self.entry_quantite.get())
            poids_unitaire = self.parser_nombre(self.entry_poids_unitaire.get())
            prix_unitaire = self.parser_nombre(self.entry_prix_unitaire.get())
            prix_vente = self.parser_nombre(self.entry_prix_vente.get())
            if quantite <= 0:
                messagebox.showwarning("Attention", "La quantité doit être > 0.")
                return
            poids_total = quantite * poids_unitaire
            prov = self.entry_provenance.get()
            date_p = self.entry_peremption.get_date().strftime('%d/%m/%Y') if self.var_has_peremption.get() else ""
            destination = self.entry_destination.get().strip()
            num_facture = self.entry_num_facture.get().strip()
            num_camion = self.entry_num_camion.get().strip()
            chauffeur = self.entry_chauffeur.get().strip()
            idx = self.index_ligne_selectionnee
            self.items_voyage[idx].update({
                'idfrs': self.fournisseur_id, 'nomfrs': prov,
                'quantite': quantite, 'poids_unitaire': poids_unitaire,
                'prix_unitaire': prix_unitaire, 'poids_total': poids_total,
                'prix_vente': prix_vente, 'date_peremption': date_p or None,
                'destination': destination, 'num_facture': num_facture,
                'num_camion': num_camion, 'chauffeur': chauffeur
            })
            item_id = self.tree.get_children()[idx]
            self.tree.item(item_id, values=(
                prov, self.article_selectionne['nomart'], self.article_selectionne['unite'],
                self.formater_nombre(quantite), self.formater_nombre(poids_unitaire),
                self.formater_nombre(prix_unitaire), self.formater_nombre(poids_total),
                self.formater_nombre(prix_vente), date_p, destination, num_facture, num_camion, chauffeur
            ))
            self.annuler_selection_ligne()
            self.calculer_total()
            messagebox.showinfo("Succès", "Ligne modifiée avec succès !")
        except ValueError:
            messagebox.showerror("Erreur", "Valeurs numériques invalides.")

    def annuler_selection_ligne(self):
        self.index_ligne_selectionnee = None
        self._reset_champs_article()
        self.btn_ajouter.configure(state="normal")
        self.btn_modifier_ligne.configure(state="disabled", text="✏️ Modif.")
        self.btn_annuler_selection.configure(state="disabled")
        self.tree.selection_remove(self.tree.selection())

    def supprimer_article(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Attention", "Sélectionnez une ligne à supprimer.")
            return
        if self.index_ligne_selectionnee is not None:
            self.annuler_selection_ligne()
        idx = self.tree.index(sel[0])
        self.tree.delete(sel[0])
        self._refresh_table_alternating_colors(self.tree)
        self.items_voyage.pop(idx)
        self.calculer_total()

    # ─────────────────────────────────────────────────────────────────────
    # Section 3 — Tableau + total + observation
    # ─────────────────────────────────────────────────────────────────────
    def _build_section_tableau(self, parent):
        card = ctk.CTkFrame(parent, fg_color=Colors.BG_CARD,
                            corner_radius=8, border_width=1, border_color=Colors.BORDER)
        card.pack(fill="both", expand=True, pady=(0, 0))

        thead = ctk.CTkFrame(card, fg_color="transparent")
        thead.pack(fill="x", padx=10, pady=(6, 4))
        ctk.CTkLabel(thead, text="📄 Marchandises du voyage",
                     font=Fonts.bold(11), text_color=Colors.MIDNIGHT).pack(side="left")

        tree_frame = ctk.CTkFrame(card, fg_color=Colors.BORDER, corner_radius=6)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=(0, 4))

        colonnes = ("Provenance", "Désignation", "Unité", "Quantité",
                    "Poids Unit.", "Prix Unit.", "Poids Total", "Prix de Vente",
                    "Péremption", "Destination", "N° Facture", "N° Camion", "Chauffeur")
        self.tree = ttk.Treeview(tree_frame, columns=colonnes, show="headings",
                                  height=8, style="iJeery.Treeview")
        self._configure_table_alternating_colors(self.tree)

        col_widths = {"Provenance": 130, "Désignation": 180, "Unité": 65, "Quantité": 80,
                      "Poids Unit.": 85, "Prix Unit.": 90, "Poids Total": 90, "Prix de Vente": 95,
                      "Péremption": 90, "Destination": 120, "N° Facture": 100, "N° Camion": 90,
                      "Chauffeur": 120}
        for col in colonnes:
            self.tree.column(col, width=col_widths.get(col, 90),
                             anchor="center" if col not in ("Provenance", "Désignation", "Destination", "Chauffeur") else "w",
                             minwidth=60)
            self.tree.heading(col, text=col)

        try:
            from treeview_sort_utils import attach_tree_sort
            attach_tree_sort(self.tree, list(colonnes), configure_columns=False)
        except ImportError:
            pass

        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        self.tree.bind('<<TreeviewSelect>>', self.on_selection_ligne)
        self.tree.bind('<Double-Button-1>', self.on_double_click_ligne)

        # Observation
        obs_wrap = ctk.CTkFrame(card, fg_color="transparent")
        obs_wrap.pack(fill="x", padx=10, pady=(0, 4))
        ctk.CTkLabel(obs_wrap, text="Observation :", font=Fonts.label(10),
                     text_color=Colors.TEXT_SECONDARY).pack(anchor="w")
        self.text_observation = ctk.CTkTextbox(obs_wrap, height=60,
                                               fg_color=Colors.BG_INPUT, border_color=Colors.BORDER,
                                               border_width=1, font=Fonts.body(11))
        self.text_observation.pack(fill="x")

        # Barre basse : supprimer + total (tonnes) + imprimer + enregistrer
        bot = ctk.CTkFrame(card, fg_color="transparent")
        bot.pack(fill="x", padx=10, pady=(0, 6))

        styled.button_danger(bot, text="Supprimer", icon="🗑️",
                             width=120, height=28, command=self.supprimer_article
                             ).pack(side="left")

        styled.button_success(bot, text="Enregistrer", icon="💾",
                              command=self.enregistrer_voyage, width=140, height=28
                              ).pack(side="right")

        styled.button_info(bot, text="Export Excel", icon="📊",
                           command=self.exporter_excel_voyage, width=140, height=28
                           ).pack(side="right", padx=(0, 8))

        self.label_total = ctk.CTkLabel(bot, text="TOTAL : 0,000 T",
                                        font=Fonts.bold(11), text_color=Colors.TEXT_MUTED)
        self.label_total.pack(side="right", padx=12)

    def calculer_total(self):
        total_kg = sum(i['poids_total'] for i in self.items_voyage)
        total_tonnes = total_kg / 1000.0
        self.label_total.configure(text=f"TOTAL : {self.formater_nombre(total_tonnes)} T")
        return total_tonnes

    # ─────────────────────────────────────────────────────────────────────
    # Génération numéro de voyage
    # ─────────────────────────────────────────────────────────────────────
    def generer_numero_voyage(self):
        conn = self.connect_db()
        if not conn:
            return
        try:
            cur = conn.cursor()
            annee = datetime.now().year
            cur.execute(
                "SELECT numero_voyage FROM logistique_voyage WHERE numero_voyage LIKE %s "
                "ORDER BY numero_voyage DESC LIMIT 1",
                (f"{annee}-VO-%",)
            )
            r = cur.fetchone()
            num = (int(r[0].split('-')[-1]) + 1) if r else 1
            ref = f"{annee}-VO-{num:05d}"
            self.entry_numero.configure(state="normal")
            self.entry_numero.delete(0, "end")
            self.entry_numero.insert(0, ref)
            self.entry_numero.configure(state="readonly")
        except Exception as e:
            messagebox.showerror("Erreur", f"Numéro de voyage : {e}")
        finally:
            if "cur" in locals():
                cur.close()
            conn.close()

    # ─────────────────────────────────────────────────────────────────────
    # Recherche / chargement d'un voyage existant
    # ─────────────────────────────────────────────────────────────────────
    def ouvrir_recherche_voyage(self):
        fen = ctk.CTkToplevel(self)
        fen.title("Charger un voyage")
        fen.geometry("900x520")
        fen.grab_set()
        Theme.apply_toplevel(fen)

        main = ctk.CTkFrame(fen, fg_color=Colors.BG_PAGE)
        main.pack(fill="both", expand=True, padx=12, pady=12)

        ctk.CTkLabel(main, text="Sélectionner un voyage",
                     font=Fonts.heading(14), text_color=Colors.MIDNIGHT).pack(pady=(0, 10))

        sf = ctk.CTkFrame(main, fg_color="transparent")
        sf.pack(fill="x", pady=(0, 8))
        ctk.CTkLabel(sf, text="🔍").pack(side="left", padx=6)
        entry_s = ctk.CTkEntry(sf, placeholder_text="Numéro ou itinéraire...", height=34,
                               fg_color=Colors.BG_INPUT, border_color=Colors.BORDER, font=Fonts.body(11))
        entry_s.pack(side="left", fill="x", expand=True, padx=4)

        tf = ctk.CTkFrame(main, fg_color=Colors.BORDER, corner_radius=8)
        tf.pack(fill="both", expand=True, pady=(0, 8))
        cols = ("ID", "Numéro", "Date création", "Itinéraire", "Statut")
        tree = ttk.Treeview(tf, columns=cols, show="headings", height=12, style="iJeery.Treeview")
        self._configure_table_alternating_colors(tree)
        tree.column("ID", width=0, stretch=False)
        tree.column("Numéro", width=130)
        tree.column("Date création", width=110)
        tree.column("Itinéraire", width=260)
        tree.column("Statut", width=110, anchor="center")
        for c in cols:
            tree.heading(c, text=c)
        tree.tag_configure('encours', background='#fff3cd')
        tree.tag_configure('cloture', background='#d4edda')
        sb = ttk.Scrollbar(tf, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=sb.set)
        tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        lbl_c = ctk.CTkLabel(main, text="", font=Fonts.small(10), text_color=Colors.TEXT_MUTED)
        lbl_c.pack(pady=(0, 4))

        def charger(filtre=""):
            for i in tree.get_children():
                tree.delete(i)
            conn = self.connect_db()
            if not conn:
                return
            try:
                cur = conn.cursor()
                q = """SELECT id, numero_voyage, date_creation, itineraire, statut
                       FROM logistique_voyage WHERE deleted=0"""
                p = []
                if filtre:
                    q += " AND (LOWER(numero_voyage) LIKE LOWER(%s) OR LOWER(itineraire) LIKE LOWER(%s))"
                    p = [f"%{filtre}%", f"%{filtre}%"]
                q += " ORDER BY date_creation DESC"
                cur.execute(q, p)
                rows = cur.fetchall()
                for r in rows:
                    ds = r[2].strftime("%d/%m/%Y") if r[2] else ""
                    label_statut = STATUT_DB_TO_LABEL.get(r[4], r[4])
                    tag = "cloture" if r[4] == "Cloture" else "encours"
                    tree.insert("", "end", values=(r[0], r[1], ds, r[3] or "", label_statut), tags=(tag,))
                lbl_c.configure(text=f"{len(rows)} voyage(s)")
            except Exception as e:
                messagebox.showerror("Erreur", str(e))
            finally:
                conn.close()

        entry_s.bind('<KeyRelease>', lambda e: charger(entry_s.get()))

        def valider():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Attention", "Sélectionnez un voyage.", parent=fen)
                return
            idvoyage = tree.item(sel[0])["values"][0]
            fen.destroy()
            self.charger_voyage(idvoyage)

        tree.bind('<Double-Button-1>', lambda e: valider())
        bf = ctk.CTkFrame(main, fg_color="transparent")
        bf.pack(fill="x")
        styled.button_danger(bf, text="Annuler", icon="❌", width=110, height=36, command=fen.destroy).pack(side="left", padx=4)
        styled.button_success(bf, text="Valider", icon="✅", width=110, height=36, command=valider).pack(side="right", padx=4)
        charger()

    def charger_voyage(self, idvoyage):
        conn = self.connect_db()
        if not conn:
            return
        try:
            cur = conn.cursor()
            cur.execute("""
                SELECT numero_voyage, vehicule_id, type_vehicule, date_creation, date_cloture,
                       statut, itineraire, observation
                FROM logistique_voyage WHERE id=%s
            """, (idvoyage,))
            v = cur.fetchone()
            if not v:
                messagebox.showerror("Erreur", "Voyage introuvable.")
                return

            self.reinitialiser_formulaire(generer_ref=False)
            self.idvoyage_charge = idvoyage
            self.mode_modification = True

            self.entry_numero.configure(state="normal")
            self.entry_numero.delete(0, "end")
            self.entry_numero.insert(0, v[0])
            self.entry_numero.configure(state="readonly")

            for cle, info in self.vehicules.items():
                if info["id"] == v[1]:
                    self.combo_vehicule.set(cle)
                    break
            self.on_vehicule_change()
            if v[2]:
                self.entry_type_vehicule.configure(state="normal")
                self.entry_type_vehicule.delete(0, "end")
                self.entry_type_vehicule.insert(0, v[2])
                self.entry_type_vehicule.configure(state="readonly")

            if v[3]:
                self.date_creation.set_date(v[3])
            if v[4]:
                self.date_cloture.set_date(v[4])

            self.combo_statut.set(STATUT_DB_TO_LABEL.get(v[5], "Encours"))
            self.entry_itineraire.delete(0, "end")
            self.entry_itineraire.insert(0, v[6] or "")

            self.text_observation.delete("1.0", "end")
            self.text_observation.insert("1.0", v[7] or "")

            cur.execute("""
                SELECT d.id, d.idfrs, f.nomfrs, d.idarticle, d.idunite, a.designation, u.designationunite,
                       d.quantite, d.poids_unitaire, d.prix_unitaire, d.poids_total, d.prix_vente,
                       d.date_peremption, d.destination, d.num_facture, d.num_camion, d.chauffeur
                FROM logistique_voyage_detail d
                LEFT JOIN tb_fournisseur f ON d.idfrs = f.idfrs
                LEFT JOIN tb_article a ON d.idarticle = a.idarticle
                LEFT JOIN tb_unite u ON d.idunite = u.idunite AND d.idarticle = u.idarticle
                WHERE d.voyage_id = %s
                ORDER BY d.id
            """, (idvoyage,))
            details = cur.fetchall()
            for d in details:
                nomfrs = d[2] or ""
                nomart = d[5] or "Article inconnu"
                unite = d[6] or ""
                date_p = d[12].strftime('%d/%m/%Y') if d[12] else None
                item = {
                    'iddetail': d[0], 'idfrs': d[1], 'nomfrs': nomfrs,
                    'idarticle': d[3], 'idunite': d[4], 'nomart': nomart, 'unite': unite,
                    'quantite': float(d[7]), 'poids_unitaire': float(d[8]),
                    'prix_unitaire': float(d[9]), 'poids_total': float(d[10]),
                    'prix_vente': float(d[11]), 'date_peremption': date_p,
                    'destination': d[13] or "", 'num_facture': d[14] or "",
                    'num_camion': d[15] or "", 'chauffeur': d[16] or ""
                }
                self.items_voyage.append(item)
                self.tree.insert("", "end", values=(
                    nomfrs, nomart, unite,
                    self.formater_nombre(item['quantite']), self.formater_nombre(item['poids_unitaire']),
                    self.formater_nombre(item['prix_unitaire']), self.formater_nombre(item['poids_total']),
                    self.formater_nombre(item['prix_vente']), date_p or "", item['destination'],
                    item['num_facture'], item['num_camion'], item['chauffeur']
                ))
            self._refresh_table_alternating_colors(self.tree)
            self.calculer_total()
        except Exception as e:
            messagebox.showerror("Erreur", f"Chargement du voyage : {e}")
        finally:
            if "cur" in locals():
                cur.close()
            conn.close()

    # ─────────────────────────────────────────────────────────────────────
    # Enregistrement
    # ─────────────────────────────────────────────────────────────────────
    def enregistrer_voyage(self):
        if self.index_ligne_selectionnee is not None:
            messagebox.showwarning("Attention", "Validez ou annulez la modification en cours.")
            return
        if not self.items_voyage:
            messagebox.showwarning("Attention", "Le voyage ne contient aucune marchandise.")
            return
        if not self.vehicule_id:
            messagebox.showwarning("Attention", "Sélectionnez un matériel de transport.")
            return

        statut_db = STATUT_LABEL_TO_DB.get(self.combo_statut.get(), "Encours")
        total_tonnes = self.calculer_total()
        date_creation = self.date_creation.get_date()
        date_cloture = self.date_cloture.get_date() if statut_db == "Cloture" else None
        itineraire = self.entry_itineraire.get().strip()
        observation = self.text_observation.get("1.0", "end").strip()
        type_vehicule = self.entry_type_vehicule.get()

        conn = self.connect_db()
        if not conn:
            return
        try:
            cur = conn.cursor()

            if self.mode_modification and self.idvoyage_charge:
                cur.execute("""
                    UPDATE logistique_voyage
                    SET vehicule_id=%s, type_vehicule=%s, date_creation=%s, date_cloture=%s,
                        statut=%s, itineraire=%s, total_poids_tonnes=%s, observation=%s, updated_at=%s
                    WHERE id=%s
                """, (self.vehicule_id, type_vehicule, date_creation, date_cloture,
                      statut_db, itineraire, total_tonnes, observation, datetime.now(),
                      self.idvoyage_charge))

                ids_existants = [i['iddetail'] for i in self.items_voyage if i['iddetail']]
                cur.execute("SELECT id FROM logistique_voyage_detail WHERE voyage_id=%s", (self.idvoyage_charge,))
                all_ids = [r[0] for r in cur.fetchall()]
                to_del = [x for x in all_ids if x not in ids_existants]
                if to_del:
                    cur.execute("DELETE FROM logistique_voyage_detail WHERE id IN %s", (tuple(to_del),))

                for item in self.items_voyage:
                    dp = self._format_date_db(item.get('date_peremption'))
                    if item['iddetail']:
                        cur.execute("""
                            UPDATE logistique_voyage_detail
                            SET idfrs=%s, idarticle=%s, idunite=%s, quantite=%s, poids_unitaire=%s,
                                prix_unitaire=%s, poids_total=%s, prix_vente=%s, date_peremption=%s,
                                destination=%s, num_facture=%s, num_camion=%s, chauffeur=%s
                            WHERE id=%s
                        """, (item['idfrs'], item['idarticle'], item['idunite'], item['quantite'],
                              item['poids_unitaire'], item['prix_unitaire'], item['poids_total'],
                              item['prix_vente'], dp, item.get('destination'), item.get('num_facture'),
                              item.get('num_camion'), item.get('chauffeur'), item['iddetail']))
                    else:
                        cur.execute("""
                            INSERT INTO logistique_voyage_detail
                                (voyage_id, idfrs, idarticle, idunite, quantite, poids_unitaire,
                                 prix_unitaire, poids_total, prix_vente, date_peremption,
                                 destination, num_facture, num_camion, chauffeur)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        """, (self.idvoyage_charge, item['idfrs'], item['idarticle'], item['idunite'],
                              item['quantite'], item['poids_unitaire'], item['prix_unitaire'],
                              item['poids_total'], item['prix_vente'], dp, item.get('destination'),
                              item.get('num_facture'), item.get('num_camion'), item.get('chauffeur')))
                conn.commit()
                messagebox.showinfo("Succès", f"Voyage {self.entry_numero.get()} modifié avec succès !")
            else:
                cur.execute("""
                    INSERT INTO logistique_voyage
                        (numero_voyage, vehicule_id, type_vehicule, date_creation, date_cloture,
                         statut, itineraire, total_poids_tonnes, observation, iduser, deleted)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,0)
                    RETURNING id
                """, (self.entry_numero.get(), self.vehicule_id, type_vehicule, date_creation,
                      date_cloture, statut_db, itineraire, total_tonnes, observation, self.iduser))
                idvoyage = cur.fetchone()[0]
                self.idvoyage_charge = idvoyage

                for item in self.items_voyage:
                    dp = self._format_date_db(item.get('date_peremption'))
                    cur.execute("""
                        INSERT INTO logistique_voyage_detail
                            (voyage_id, idfrs, idarticle, idunite, quantite, poids_unitaire,
                             prix_unitaire, poids_total, prix_vente, date_peremption,
                             destination, num_facture, num_camion, chauffeur)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """, (idvoyage, item['idfrs'], item['idarticle'], item['idunite'],
                          item['quantite'], item['poids_unitaire'], item['prix_unitaire'],
                          item['poids_total'], item['prix_vente'], dp, item.get('destination'),
                          item.get('num_facture'), item.get('num_camion'), item.get('chauffeur')))
                conn.commit()
                messagebox.showinfo("Succès", "Voyage enregistré avec succès !")

            exporter = YesNoDialog(
                "Export",
                "Le voyage a été enregistré.\n\nVoulez-vous exporter le bon de voyage vers Excel ?"
            ).result
            if exporter:
                self.exporter_excel_voyage()

            self.mode_modification = False
            self.reinitialiser_formulaire()

        except Exception as e:
            conn.rollback()
            messagebox.showerror("Erreur", f"Enregistrement : {e}")
        finally:
            if "cur" in locals():
                cur.close()
            conn.close()

    # ─────────────────────────────────────────────────────────────────────
    # Réinitialisation / Nouveau
    # ─────────────────────────────────────────────────────────────────────
    def reinitialiser_formulaire(self, generer_ref=True):
        if generer_ref:
            self.generer_numero_voyage()
        self.items_voyage.clear()
        self.index_ligne_selectionnee = None
        self.article_selectionne = None
        self.idvoyage_charge = None
        for i in self.tree.get_children():
            self.tree.delete(i)
        self._refresh_table_alternating_colors(self.tree)
        self._reset_champs_article()
        self.btn_ajouter.configure(state="normal")
        self.btn_modifier_ligne.configure(state="disabled", text="✏️ Modif.")
        self.btn_annuler_selection.configure(state="disabled")
        self.entry_itineraire.delete(0, "end")
        self.text_observation.delete("1.0", "end")
        self.combo_statut.set("Encours")
        self.date_creation.set_date(datetime.now())
        self.date_cloture.set_date(datetime.now())
        self.calculer_total()

    def nouveau_voyage(self):
        self.reinitialiser_formulaire()
        self.mode_modification = False
        self.idvoyage_charge = None

    # ─────────────────────────────────────────────────────────────────────
    # EXPORT EXCEL — remplace l'ancien aperçu PDF
    # ─────────────────────────────────────────────────────────────────────
    def exporter_excel_voyage(self):
        if not self.items_voyage:
            messagebox.showwarning("Attention", "Aucune marchandise à exporter.")
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror(
                "Erreur",
                "Le module 'openpyxl' est requis pour l'export Excel.\n"
                "Installez-le avec : pip install openpyxl"
            )
            return

        numero = self.entry_numero.get()
        chemin = filedialog.asksaveasfilename(
            title="Exporter le voyage en Excel",
            defaultextension=".xlsx",
            initialfile=f"Voyage_{numero}.xlsx",
            filetypes=[("Classeur Excel", "*.xlsx")]
        )
        if not chemin:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Voyage"

        bold = Font(bold=True)
        title_font = Font(bold=True, size=14, color="2C3E50")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("A1:F1")
        ws["A1"] = f"Fiche de Voyage — {numero}"
        ws["A1"].font = title_font

        date_cl = (self.date_cloture.get_date().strftime('%d/%m/%Y')
                   if self.combo_statut.get() == "Clôturé" else "—")
        infos = [
            ("Voyage N°", numero),
            ("Matériel de transport", self.combo_vehicule.get()),
            ("Type véhicule", self.entry_type_vehicule.get()),
            ("Date de création", self.date_creation.get_date().strftime('%d/%m/%Y')),
            ("Date de clôture", date_cl),
            ("Statut", self.combo_statut.get()),
            ("Itinéraire", self.entry_itineraire.get()),
        ]
        row = 3
        for label, value in infos:
            ws.cell(row=row, column=1, value=label).font = bold
            ws.cell(row=row, column=2, value=value)
            row += 1

        row += 1
        entetes = ["N°", "Provenance", "Désignation", "Unité", "Quantité", "Poids Unit.",
                   "Prix Unit.", "Poids Total", "Prix de Vente", "Péremption",
                   "Destination", "N° Facture", "N° Camion", "Chauffeur"]
        header_row = row
        for col, txt in enumerate(entetes, start=1):
            c = ws.cell(row=header_row, column=col, value=txt)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center

        total_poids_kg = 0
        r = header_row + 1
        for i, item in enumerate(self.items_voyage, start=1):
            total_poids_kg += item['poids_total']
            ws.cell(row=r, column=1, value=i)
            ws.cell(row=r, column=2, value=item.get('nomfrs') or "")
            ws.cell(row=r, column=3, value=item.get('nomart') or "")
            ws.cell(row=r, column=4, value=item.get('unite') or "")
            ws.cell(row=r, column=5, value=item['quantite'])
            ws.cell(row=r, column=6, value=item['poids_unitaire'])
            ws.cell(row=r, column=7, value=item['prix_unitaire'])
            ws.cell(row=r, column=8, value=item['poids_total'])
            ws.cell(row=r, column=9, value=item['prix_vente'])
            ws.cell(row=r, column=10, value=item.get('date_peremption') or "")
            ws.cell(row=r, column=11, value=item.get('destination') or "")
            ws.cell(row=r, column=12, value=item.get('num_facture') or "")
            ws.cell(row=r, column=13, value=item.get('num_camion') or "")
            ws.cell(row=r, column=14, value=item.get('chauffeur') or "")
            r += 1

        total_tonnes = total_poids_kg / 1000.0
        ws.cell(row=r, column=7, value="TOTAL (Tonnes)").font = bold
        ws.cell(row=r, column=8, value=round(total_tonnes, 3)).font = bold
        r += 2

        ws.cell(row=r, column=1, value="Observation :").font = bold
        ws.cell(row=r, column=2, value=self.text_observation.get("1.0", "end").strip())

        largeurs = [5, 20, 25, 10, 10, 12, 12, 12, 12, 14, 16, 14, 12, 16]
        for i, l in enumerate(largeurs, start=1):
            ws.column_dimensions[get_column_letter(i)].width = l

        try:
            wb.save(chemin)
        except Exception as e:
            messagebox.showerror("Erreur", f"Export Excel : {e}")
            return

        if messagebox.askyesno("Export réussi", f"Fichier enregistré :\n{chemin}\n\nVoulez-vous l'ouvrir maintenant ?"):
            try:
                os.startfile(chemin)
            except Exception:
                pass



# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    from app_theme import init_theme
    init_theme()
    app = ctk.CTk()
    app.geometry("1360x860")
    app.title("Voyage — iJeery")
    Theme.apply(app)
    page = PageVoyage(app, iduser=1)
    page.pack(fill="both", expand=True)
    app.mainloop()
